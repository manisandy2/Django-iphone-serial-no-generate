from django.shortcuts import render
import sqlite3
from .models import Generate_Serial,Product_List
from django.shortcuts import HttpResponse,HttpResponseRedirect
from .forms import LogInForm,ProductListForm
from django.db.models import Q
from django.contrib.auth import login,logout,authenticate
from django.contrib import messages
import datetime
import sqlite3
from openpyxl import load_workbook


def product_data_list():
    wb = load_workbook(r'C:\Users\admin\Desktop\Serial\myweb\excel\Ingram Serial Num.xlsx')
    ws = wb.active
    col = ws.max_column
    row = ws.max_row
    con = sqlite3.connect(r'C:\Users\admin\Desktop\Serial\myweb\db.sqlite3')
    cur = con.cursor()

    product_data = []
    for r in range(2, row+1):
        mpn = ws.cell(r, 1).value
        ingram_code = ws.cell(r, 2).value
        apx_model = ws.cell(r, 3).value
        brand = ws.cell(r, 4).value
        product = ws.cell(r, 5).value
        item_code = ws.cell(r, 6).value
        id_value = r - 1
        product_data = [(id_value, mpn,ingram_code,apx_model,brand,product,item_code)]
        cur.executemany("INSERT INTO generate VALUES(?,?,?,?,?,?,?)",product_data)
        con.commit()
    con.close()


def product(request):
    return render(request, "index.html")


def login_page(request):
    if not request.user.is_authenticated:
        if request.method == "POST":
            fm = LogInForm(request=request,data=request.POST)
            if fm.is_valid():
                uname = fm.cleaned_data['username']
                upass = fm.cleaned_data['password']
                user = authenticate(username=uname,password=upass)
                if user is not None:
                    login(request,user)
                    messages.success(request,"Logged in Successfully")
                    return HttpResponseRedirect('/')
        fm = LogInForm()
        return render(request, 'login.html', {'form': fm})
    else:
        return HttpResponseRedirect('/dashboard/')


def logout_page(request):
    logout(request)
    return HttpResponseRedirect('/')


def dashboard_page(request):
    if request.user.is_authenticated:
        product_all = Product_List.objects.all().order_by('-pk')
        gs = Generate_Serial.objects.all()
        return render(request, "dashboard.html", {"product": product_all})
    else:
        return HttpResponseRedirect("/login/")


def dashboard_filter(request,itemcode):
    if request.user.is_authenticated:
        gs = Generate_Serial.objects.filter(item_code=itemcode)
        return render(request, "dashboard_filter.html", {"gs": gs})
    else:
        return HttpResponseRedirect("/login/")


def generate_page(request):
    if request.user.is_authenticated:
        if request.method == "POST":
            plf = ProductListForm(request.POST)
            if plf.is_valid():
                i_name = plf.cleaned_data['list_of_name']
                i_search = plf.cleaned_data['list_of_code']
                # if str(i_search[0,1]) == "M":
                if str(i_search[0:1]) == "M":
                    i_name = "mpn"
                elif str(i_search[0:1]) == "G":
                    i_name = "ingram_code"
                else:
                    i_name = "item_code"
                i_range = plf.cleaned_data['list_of_range']
                i_name_create = plf.cleaned_data['name_invoice_create']
                i_date = datetime.datetime.now()
                i_invoice = plf.cleaned_data['invoice']
                # print(i_name_create)
                # print(i_invoice)
                # print(i_name)
                data_sh = Generate_Serial.objects.get(
                    Q(mpn=i_search) | Q(ingram_code=i_search) | Q(item_code=i_search))
                pls = Product_List.objects.all().filter(list_of_itemcode=data_sh.item_code).order_by('-serial')
                r = 1
                for p in pls:
                    if r == 1:
                        # print(p.name, p.date, p.serial)
                        p_serial = p.serial
                        r = r + 1
                        break
                all_product = str(data_sh.item_code) + str(data_sh.mpn[0:6])
                plf.serial = all_product
                if request.method == "POST":
                    for r in range(1, i_range+1):
                        pl = Product_List.objects.create()
                        if not pls:
                            pl.serial = plf.serial + '{:d}'.format(r).zfill(4)
                        else:
                            num = int(p_serial[11:len(p_serial)]) + r
                            pl.serial = plf.serial + '{:d}'.format(num).zfill(4)
                        pl.name = request.user
                        pl.list_of_range = i_range
                        pl.list_of_name = i_name
                        pl.list_of_code = i_search
                        pl.list_of_itemcode = data_sh.item_code
                        pl.name_invoice_create = i_name_create
                        pl.invoice = i_invoice
                        pl.save()
                    if i_name == "item_code":
                        pls = Product_List.objects.all().filter(list_of_code=i_search).order_by('-serial')[0:i_range]
                        gs = Generate_Serial.objects.get(item_code=i_search)
                    elif i_name == "ingram_code":
                        pls = Product_List.objects.all().filter(list_of_code=i_search).order_by('-serial')[0:i_range]
                        gs = Generate_Serial.objects.get(ingram_code=i_search)
                    elif i_name == "mpn":
                        pls = Product_List.objects.all().filter(list_of_code=i_search).order_by('-serial')[0:i_range]
                        gs = Generate_Serial.objects.get(mpn=i_search)
                return render(request, "generate_page.html", {"pl": pls, "gs": gs})
        else:
            plf = ProductListForm()
        return render(request, "generate_page.html", {"plf": plf})
    else:
        return HttpResponseRedirect("/login/")

