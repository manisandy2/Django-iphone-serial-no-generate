import sqlite3
from openpyxl import load_workbook


def product_data_list():
    wb = load_workbook(r'C:\Users\admin\Desktop\Serial\myweb\excel\Ingram Serial Num.xlsx')
    ws = wb.active
    print(ws.title)
    col = ws.max_column
    row = ws.max_row
    print(col)
    print(row)
    con = sqlite3.connect(r'C:\Users\admin\Desktop\Serial\myweb\db.sqlite3')
    print("Opened database successfully")
    cur = con.cursor()

    product_data = []
    for r in range(2, row+1):
        mpn = ws.cell(r, 1).value
        ingram_code = ws.cell(r, 2).value
        apx_model = ws.cell(r, 3).value
        brand = ws.cell(r, 4).value
        product = ws.cell(r, 5).value
        item_code = ws.cell(r, 6).value
        id_value = r - 1
        product_data = [(id_value, mpn,ingram_code,item_code,apx_model,brand,product)]
        # cur.executemany("INSERT INTO generate VALUES(?,?,?,?,?,?,?)", (1, 'BUY', 'RHAT', 'ZXc', 'adsfad', 'asdf', 'jk'))
        cur.executemany("INSERT INTO generate VALUES(?,?,?,?,?,?,?)",product_data)
        con.commit()
    con.close()



product_data_list()
